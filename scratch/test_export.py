import requests
import os
from dotenv import load_dotenv

load_dotenv()

BASE_URL = f"http://localhost:{os.getenv('PORT', 5001)}"

def test_export():
    # 1. Login
    login_data = {"username": "kartik@gmail.com", "password": "Kartik@123"}
    resp = requests.post(f"{BASE_URL}/auth/login", json=login_data)
    if resp.status_code != 200:
        print(f"Login failed: {resp.json()}")
        return

    token = resp.json().get("token")
    headers = {"Authorization": f"Bearer {token}"}

    # 2. Trigger Export
    print("Triggering timesheet export...")
    export_resp = requests.get(f"{BASE_URL}/timesheets/export", headers=headers)
    
    if export_resp.status_code == 200:
        content_type = export_resp.headers.get("Content-Type")
        print(f"Success! Content-Type: {content_type}")
        
        # Save file to scratch for verification
        filename = "scratch/test_export_result.xlsx"
        with open(filename, "wb") as f:
            f.write(export_resp.content)
        print(f"File saved to {filename}")
        
        # Verify with openpyxl
        import openpyxl
        wb = openpyxl.load_workbook(filename)
        print(f"Worksheets found: {wb.sheetnames}")
        if "Timesheet Data" in wb.sheetnames:
            print("Verified: 'Timesheet Data' sheet exists.")
        if "Summary" in wb.sheetnames:
            print("Verified: 'Summary' sheet exists.")
    else:
        print(f"Export failed with status {export_resp.status_code}: {export_resp.text}")

if __name__ == "__main__":
    # Check if server is running
    try:
        requests.get(BASE_URL)
    except:
        print(f"Server not running at {BASE_URL}. Please start it first.")
    else:
        test_export()
