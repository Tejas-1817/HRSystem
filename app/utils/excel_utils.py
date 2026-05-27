import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from app.utils.display_name_service import get_clean_name

def generate_timesheet_excel(employee_name, timesheets, start_date_str=None, end_date_str=None):
    """
    Generates a professional, enterprise-grade Excel workbook for timesheet data.
    Adheres to modern SaaS reporting standards:
    - Clean typography (Segoe UI)
    - Corporate styling (Dark Slate/Navy headers, soft zebra striping)
    - Metadata/Title Block at the top of sheets
    - Explicit gridlines visibility
    - Summary dashboard sheet with dynamic counts and project/billability statistics
    """
    wb = Workbook()
    
    # --- Color Palette & Font Config ---
    FONT_NAME = 'Segoe UI'
    
    # Styles
    title_font = Font(name=FONT_NAME, size=16, bold=True, color="1F4E78")
    section_font = Font(name=FONT_NAME, size=12, bold=True, color="2C3E50")
    header_font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=FONT_NAME, size=10)
    bold_data_font = Font(name=FONT_NAME, size=10, bold=True)
    metadata_label_font = Font(name=FONT_NAME, size=10, bold=True, color="595959")
    metadata_value_font = Font(name=FONT_NAME, size=10, color="000000")
    
    # Fills
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    summary_header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    zebra_even_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    zebra_odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Borders & Alignments
    thin_side = Side(style='thin', color="D1D5DB")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    thick_bottom_side = Side(style='medium', color="1F4E78")
    double_bottom_side = Side(style='double', color="000000")
    total_border = Border(top=thin_side, bottom=double_bottom_side)
    
    center_aligned = Alignment(horizontal="center", vertical="center")
    left_aligned = Alignment(horizontal="left", vertical="center")
    right_aligned = Alignment(horizontal="right", vertical="center")
    
    # ---------------------------------------------------------
    # 1. Main Data Sheet
    # ---------------------------------------------------------
    ws = wb.active
    ws.title = "Timesheet Data"
    ws.views.sheetView[0].showGridLines = True
    
    # Add Metadata Title Block
    ws.cell(row=1, column=1, value="HR Management System - Timesheet Report").font = title_font
    
    # Metadata Fields
    metadata = [
        ("Team Member:", get_clean_name(employee_name)),
        ("Report Period:", f"{start_date_str} to {end_date_str}" if (start_date_str or end_date_str) else "All Records"),
        ("Export Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    ]
    
    for i, (label, val) in enumerate(metadata, 3):
        lbl_cell = ws.cell(row=i, column=1, value=label)
        lbl_cell.font = metadata_label_font
        lbl_cell.alignment = left_aligned
        
        val_cell = ws.cell(row=i, column=2, value=val)
        val_cell.font = metadata_value_font
        val_cell.alignment = left_aligned
        
    # Table Header (Row 7)
    headers = [
        "Team Member", "Project Name", "Manager Name", "Task", "Description", 
        "Date", "Hours Worked", "Billable Status"
    ]
    
    header_row_idx = 7
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row_idx, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned
        cell.border = border
    
    # Add Data Rows
    total_hours = 0
    project_hours = {}
    billable_hours = {"Billable": 0, "Non-Billable": 0}
    unique_team_members = set()
    
    data_start_row = 8
    for idx, ts in enumerate(timesheets):
        row_num = data_start_row + idx
        curr_member = get_clean_name(ts.get("employee_name", employee_name))
        project = ts.get("project", "N/A")
        hours = float(ts.get("hours", 0))
        is_billable = "Billable" if ts.get("is_billable") or ts.get("billable") else "Non-Billable"
        
        unique_team_members.add(curr_member)
        
        row_data = [
            curr_member,
            project,
            ts.get("manager_name", "N/A"),
            ts.get("task", "N/A"),
            ts.get("description", ""),
            ts.get("start_date").isoformat() if hasattr(ts.get("start_date"), "isoformat") else str(ts.get("start_date")),
            hours,
            is_billable
        ]
        
        # Apply zebra striping based on index
        row_fill = zebra_even_fill if idx % 2 == 0 else zebra_odd_fill
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = data_font
            cell.border = border
            cell.fill = row_fill
            
            # Alignments & Formatting
            if col_num == 6:  # Date
                cell.alignment = center_aligned
            elif col_num == 7:  # Hours column
                cell.alignment = right_aligned
                cell.number_format = '0.00'
            elif col_num == 8:  # Billable Status
                cell.alignment = center_aligned
            else:
                cell.alignment = left_aligned
                
        # Aggregation for summary
        total_hours += hours
        project_hours[project] = project_hours.get(project, 0) + hours
        billable_hours[is_billable] += hours

    # Total Row at the bottom of data
    total_row_idx = data_start_row + len(timesheets)
    ws.cell(row=total_row_idx, column=6, value="Total").font = bold_data_font
    ws.cell(row=total_row_idx, column=6).alignment = right_aligned
    ws.cell(row=total_row_idx, column=6).border = total_border
    
    total_hours_cell = ws.cell(row=total_row_idx, column=7, value=total_hours)
    total_hours_cell.font = bold_data_font
    total_hours_cell.alignment = right_aligned
    total_hours_cell.number_format = '0.00'
    total_hours_cell.border = total_border
    
    # Border styling for remaining columns in Total row
    for col_num in range(1, len(headers) + 1):
        if col_num not in (6, 7):
            ws.cell(row=total_row_idx, column=col_num).border = Border(top=thin_side)

    # Auto-adjust column widths dynamically
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            # Skip title row and metadata rows when calculating widths
            if cell.row < header_row_idx:
                continue
            try:
                if cell.value:
                    # Formatting numbers adds string length
                    val_str = f"{cell.value:.2f}" if isinstance(cell.value, (int, float)) and cell.column == 7 else str(cell.value)
                    if len(val_str) > max_length:
                        max_length = len(val_str)
            except:
                pass
        ws.column_dimensions[column_letter].width = max(max_length + 4, 12)

    # ---------------------------------------------------------
    # 2. Summary Sheet
    # ---------------------------------------------------------
    summary_ws = wb.create_sheet(title="Summary")
    summary_ws.views.sheetView[0].showGridLines = True
    
    # Title
    summary_ws.cell(row=1, column=1, value="HR Management System - Timesheet Summary").font = title_font
    
    # Period details in Summary
    summary_ws.cell(row=3, column=1, value="Report Period:").font = metadata_label_font
    summary_ws.cell(row=3, column=2, value=metadata[1][1]).font = metadata_value_font
    
    # Summary KPI / Metric Table
    summary_ws.cell(row=5, column=1, value="Key Performance Indicators").font = section_font
    
    kpis = [
        ("Total Team Members", len(unique_team_members)),
        ("Grand Total Hours", total_hours)
    ]
    
    # KPI Headers
    summary_ws.cell(row=6, column=1, value="Metric").font = header_font
    summary_ws.cell(row=6, column=1).fill = summary_header_fill
    summary_ws.cell(row=6, column=1).border = border
    summary_ws.cell(row=6, column=2, value="Value").font = header_font
    summary_ws.cell(row=6, column=2).fill = summary_header_fill
    summary_ws.cell(row=6, column=2).alignment = right_aligned
    summary_ws.cell(row=6, column=2).border = border
    
    for idx, (metric, val) in enumerate(kpis):
        r_idx = 7 + idx
        c1 = summary_ws.cell(row=r_idx, column=1, value=metric)
        c1.font = data_font
        c1.border = border
        
        c2 = summary_ws.cell(row=r_idx, column=2, value=val)
        c2.font = bold_data_font
        c2.border = border
        if isinstance(val, (int, float)):
            c2.alignment = right_aligned
            if metric == "Grand Total Hours":
                c2.number_format = '0.00'
        else:
            c2.alignment = left_aligned
            
    # Project Breakdown
    start_r = 11
    summary_ws.cell(row=start_r, column=1, value="Project Breakdown").font = section_font
    
    # Headers
    p_headers = ["Project Name", "Hours Logged", "% of Total"]
    for col_num, h_val in enumerate(p_headers, 1):
        c = summary_ws.cell(row=start_r + 1, column=col_num, value=h_val)
        c.font = header_font
        c.fill = summary_header_fill
        c.border = border
        if col_num in (2, 3):
            c.alignment = right_aligned
        else:
            c.alignment = left_aligned
            
    row_offset = start_r + 2
    for idx, (proj, hrs) in enumerate(project_hours.items()):
        curr_row = row_offset + idx
        pct = (hrs / total_hours) if total_hours > 0 else 0
        
        c1 = summary_ws.cell(row=curr_row, column=1, value=proj)
        c1.font = data_font
        c1.border = border
        
        c2 = summary_ws.cell(row=curr_row, column=2, value=hrs)
        c2.font = data_font
        c2.alignment = right_aligned
        c2.number_format = '0.00'
        c2.border = border
        
        c3 = summary_ws.cell(row=curr_row, column=3, value=pct)
        c3.font = data_font
        c3.alignment = right_aligned
        c3.number_format = '0.0%'
        c3.border = border
        
    # Project Total Row
    proj_total_row = row_offset + len(project_hours)
    summary_ws.cell(row=proj_total_row, column=1, value="Total").font = bold_data_font
    summary_ws.cell(row=proj_total_row, column=1).border = total_border
    
    c_hrs_total = summary_ws.cell(row=proj_total_row, column=2, value=total_hours)
    c_hrs_total.font = bold_data_font
    c_hrs_total.alignment = right_aligned
    c_hrs_total.number_format = '0.00'
    c_hrs_total.border = total_border
    
    c_pct_total = summary_ws.cell(row=proj_total_row, column=3, value=1.0 if total_hours > 0 else 0.0)
    c_pct_total.font = bold_data_font
    c_pct_total.alignment = right_aligned
    c_pct_total.number_format = '0.0%'
    c_pct_total.border = total_border
    
    # Billability Breakdown
    bill_start_r = proj_total_row + 3
    summary_ws.cell(row=bill_start_r, column=1, value="Billability Breakdown").font = section_font
    
    b_headers = ["Billability Status", "Hours Logged", "% of Total"]
    for col_num, h_val in enumerate(b_headers, 1):
        c = summary_ws.cell(row=bill_start_r + 1, column=col_num, value=h_val)
        c.font = header_font
        c.fill = summary_header_fill
        c.border = border
        if col_num in (2, 3):
            c.alignment = right_aligned
        else:
            c.alignment = left_aligned
            
    b_row_offset = bill_start_r + 2
    for idx, (status, hrs) in enumerate(billable_hours.items()):
        curr_row = b_row_offset + idx
        pct = (hrs / total_hours) if total_hours > 0 else 0
        
        c1 = summary_ws.cell(row=curr_row, column=1, value=status)
        c1.font = data_font
        c1.border = border
        
        c2 = summary_ws.cell(row=curr_row, column=2, value=hrs)
        c2.font = data_font
        c2.alignment = right_aligned
        c2.number_format = '0.00'
        c2.border = border
        
        c3 = summary_ws.cell(row=curr_row, column=3, value=pct)
        c3.font = data_font
        c3.alignment = right_aligned
        c3.number_format = '0.0%'
        c3.border = border
        
    # Billable Total Row
    bill_total_row = b_row_offset + len(billable_hours)
    summary_ws.cell(row=bill_total_row, column=1, value="Total").font = bold_data_font
    summary_ws.cell(row=bill_total_row, column=1).border = total_border
    
    cb_hrs_total = summary_ws.cell(row=bill_total_row, column=2, value=total_hours)
    cb_hrs_total.font = bold_data_font
    cb_hrs_total.alignment = right_aligned
    cb_hrs_total.number_format = '0.00'
    cb_hrs_total.border = total_border
    
    cb_pct_total = summary_ws.cell(row=bill_total_row, column=3, value=1.0 if total_hours > 0 else 0.0)
    cb_pct_total.font = bold_data_font
    cb_pct_total.alignment = right_aligned
    cb_pct_total.number_format = '0.0%'
    cb_pct_total.border = total_border
    
    # Auto-adjust column widths for Summary
    for col in summary_ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            # Ignore title block row when calculating width
            if cell.row == 1:
                continue
            try:
                if cell.value:
                    val_str = str(cell.value)
                    if len(val_str) > max_length:
                        max_length = len(val_str)
            except:
                pass
        summary_ws.column_dimensions[column_letter].width = max(max_length + 5, 18)

    # Finalize
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
